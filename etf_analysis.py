import openpyxl
import os
from setting import  get_file_path, get_file_name

file_path = get_file_path()
file_name = get_file_name()

# 切換到指定路徑
os.chdir(file_path)

# 讀進Excel檔案
wb = openpyxl.load_workbook(file_name)

# 取的Excel的第一個工作表
sheet = wb.worksheets[0]

etf_all = dict()

# 彙整全部的ETF清單
for columnNum in range(1, sheet.max_column + 1, 3):
    for rowNum in range(3, sheet.max_row + 1):
        if (sheet.cell(rowNum, columnNum).value == None):
            break
        if (etf_all.get(sheet.cell(rowNum, columnNum).value) == None):
            etf_all[sheet.cell(rowNum, columnNum).value] = {
                'name' : sheet.cell(rowNum, columnNum + 1).value, 
                'content' : [sheet.cell(1, columnNum).value]
                }
        else:
            etf_all.get(sheet.cell(rowNum, columnNum).value)['content'].append(sheet.cell(1, columnNum).value)

sorted_list = sorted(etf_all.items(), key=lambda x:len(x[1]['content']), reverse=True)

# 輸出的結果
new_sheet = wb.create_sheet('result')
row = 1
column = 1
for t in sorted_list:
    new_sheet.cell(row, column).value = t[0]
    new_sheet.cell(row, column + 1).value = t[1]['name']
    new_sheet.cell(row, column + 2).value = len(t[1]['content'])
    new_sheet.cell(row, column + 3).value = ','.join(str(etf_id) for etf_id in t[1]['content'])
    row = row + 1

# 存檔
wb.save(file_name)
